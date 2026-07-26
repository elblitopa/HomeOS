"""Genera el archivo de Excel del presupuesto."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MONEY = '"$"#,##0.00'
HEADER_FILL = PatternFill("solid", fgColor="2383E2")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)


def _write_table(ws, headers, rows, money_cols=()):
    start = ws.max_row + 1 if ws.max_row > 1 else ws.max_row
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
        for index in money_cols:
            ws.cell(row=ws.max_row, column=index + 1).number_format = MONEY
    return start


def _autofit(ws):
    for column in ws.columns:
        width = max((len(str(c.value)) for c in column if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(42, width + 4)


def build_budget_xlsx(data: dict) -> bytes:
    wb = Workbook()

    # --- Resumen ---
    ws = wb.active
    ws.title = "Resumen"
    ws.append([f"Presupuesto mensual · {data['month_label']}"])
    ws["A1"].font = TITLE_FONT
    ws.append([])
    _write_table(
        ws,
        ["Concepto", "Monto (MXN)"],
        [
            ["Suscripciones", data["totals"]["subscriptions"]],
            ["Pagos recurrentes", data["totals"]["recurring"]],
            ["Metas (ahorro del mes)", data["totals"]["goals"]],
            ["Total comprometido al mes", data["totals"]["commitments"]],
            [],
            ["Ingreso esperado", data["totals"]["expected_income"]],
            ["Ingreso real del mes", data["totals"]["actual_income"]],
            ["Diferencia vs esperado", data["totals"]["balance_expected"]],
            ["Diferencia vs real", data["totals"]["balance_actual"]],
        ],
        money_cols=(1,),
    )
    _autofit(ws)

    # --- Compromisos ---
    ws = wb.create_sheet("Compromisos")
    _write_table(
        ws,
        ["Tipo", "Nombre", "Monto", "Divisa", "Periodo", "Equivale al mes (MXN)", "Cuenta", "Próximo pago", "Nota"],
        [
            [
                c["kind"], c["name"], c["amount"], c["currency"], c["period"],
                c["monthly_mxn"], c["account"] or "—",
                (c["next_due"] or "")[:10], c["note"] or "",
            ]
            for c in data["commitments"]
        ],
        money_cols=(2, 5),
    )
    _autofit(ws)

    # --- Ingresos ---
    ws = wb.create_sheet("Ingresos")
    _write_table(
        ws,
        ["Cuenta", "Divisa", "Esperado mensual", "Esperado (MXN)", "Real del mes (MXN)"],
        [
            [i["account"], i["currency"], i["expected"], i["expected_mxn"], i["actual_mxn"]]
            for i in data["income"]
        ],
        money_cols=(2, 3, 4),
    )
    _autofit(ws)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
